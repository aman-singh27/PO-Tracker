import csv, io
from datetime import date
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.core.exceptions import ValidationError as DjangoValidationError
from django.db import IntegrityError, transaction
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework import status
from .models import PurchaseOrder,POLineItem,ImportReviewItem,Challan,ChallanAllocation,Bill,Payment,Client,Site,POAttachment,LegalEntity,TrackerSettings
from django.views.decorators.csrf import csrf_exempt
from .permissions import IsTrackerUser,CanEditPO,CanShortClose,CanRecordMoney,IsAdmin,get_role
from .serializers import POSerializer,SimplePOSerializer,LineSerializer,ChallanSerializer,BillSerializer,PaymentSerializer
from .services import create_po,revise_po,short_close_line,create_challan,create_bill,create_payment,allocate_bill,validate_bill_number
from .selectors import search, line_status, po_totals, pending_lines, dashboard_snapshot

def _validation(exc): return Response(getattr(exc,'message_dict',{'detail':exc.messages if hasattr(exc,'messages') else str(exc)}),status=400)
@api_view(['POST'])
@permission_classes([AllowAny])
@csrf_exempt
def login_view(request):
    user=authenticate(request,username=request.data.get('email',''),password=request.data.get('password',''))
    if not user: return Response({'detail':'Invalid email or password.'},status=401)
    if not get_role(user): return Response({'detail':'This account has no tracker role.'},status=403)
    login(request,user); return Response({'user':{'id':user.id,'email':user.email,'name':user.get_full_name()},'role':get_role(user),'force_password_change':user.tracker_role.force_password_change})
@api_view(['POST'])
def logout_view(request): logout(request); return Response(status=204)
@api_view(['GET'])
@permission_classes([IsTrackerUser])
def me_view(request):
    u=request.user; return Response({'user':{'id':u.id,'email':u.email,'name':u.get_full_name()},'role':get_role(u),'force_password_change':u.tracker_role.force_password_change})
@api_view(['POST'])
@permission_classes([IsTrackerUser])
def change_password(request):
    password=request.data.get('password','')
    if len(password)<8:return Response({'password':['Must be at least 8 characters.']},status=400)
    request.user.set_password(password); request.user.save(); request.user.tracker_role.force_password_change=False; request.user.tracker_role.save(update_fields=['force_password_change']); return Response(status=204)
@api_view(['POST'])
@permission_classes([IsAdmin])
def reset_password(request,user_id):
    password=request.data.get('password','')
    if len(password)<8:return Response({'password':['Must be at least 8 characters.']},status=400)
    user=get_object_or_404(get_user_model(),pk=user_id); user.set_password(password); user.save(); user.tracker_role.force_password_change=True; user.tracker_role.save(); return Response(status=204)
@api_view(['GET','POST'])
@permission_classes([IsTrackerUser])
def po_list(request):
    if request.method=='GET':
        qs=PurchaseOrder.objects.filter(is_deleted=False).select_related('client','site').order_by('-updated_at')
        for field in ('client','site','status','needs_review'):
            if request.query_params.get(field) is not None: qs=qs.filter(**{field:request.query_params[field]})
        return Response(SimplePOSerializer(qs,many=True).data)
    serializer=POSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    try: po=create_po(data=dict(serializer.validated_data),actor=request.user)
    except DjangoValidationError as exc: return _validation(exc)
    return Response(POSerializer(po).data,status=201)
@api_view(['GET','PATCH','DELETE'])
@permission_classes([IsTrackerUser])
def po_detail(request,pk):
    po=get_object_or_404(PurchaseOrder.objects.prefetch_related('lines__challan_allocations','lines__bill_allocations'),pk=pk,is_deleted=False)
    if request.method=='GET':
        data=POSerializer(po).data; data['totals']=po_totals(po); return Response(data)
    if request.method=='DELETE': po.is_deleted=True; po.updated_by=request.user; po.save(); return Response(status=204)
    if request.data.get('updated_at') and request.data['updated_at'] != po.updated_at.isoformat().replace('+00:00','Z'): return Response({'detail':'This PO was updated by another user. Reload and retry.'},status=409)
    serializer=POSerializer(po,data=request.data,partial=True); serializer.is_valid(raise_exception=True); serializer.save(updated_by=request.user); return Response(serializer.data)
@api_view(['GET','POST'])
@permission_classes([IsTrackerUser])
def po_attachments(request,pk):
    po=get_object_or_404(PurchaseOrder,pk=pk,is_deleted=False)
    if request.method=='GET': return Response([{'id':item.id,'label':item.label,'file':item.file.url,'created_at':item.created_at} for item in po.attachments.all().order_by('-created_at')])
    upload=request.FILES.get('file')
    if not upload:return Response({'file':['Choose a file to upload.']},status=400)
    if upload.size > 20*1024*1024:return Response({'file':['Files must be 20 MB or smaller.']},status=400)
    attachment=POAttachment.objects.create(po=po,file=upload,label=request.data.get('label','') or upload.name,uploaded_by=request.user)
    return Response({'id':attachment.id,'label':attachment.label,'file':attachment.file.url,'created_at':attachment.created_at},status=201)
@api_view(['GET','POST'])
@permission_classes([IsTrackerUser])
def po_activity(request,pk):
    po=get_object_or_404(PurchaseOrder,pk=pk,is_deleted=False)
    if request.method=='GET':
        bills=Bill.objects.filter(is_deleted=False,allocations__line_item__po=po).distinct().order_by('-bill_date')
        return Response([{'id':bill.id,'bill_number':bill.bill_number,'bill_date':bill.bill_date,'total_amount':str(bill.total_amount),'ariba_state':bill.ariba_state} for bill in bills])
    kind=request.data.get('kind'); entry_date=request.data.get('date') or date.today(); number=request.data.get('number','').strip()
    if kind=='payment':
        bill_id=request.data.get('bill_id')
        if not bill_id:return Response({'bill_id':['Select the bill that received this payment.']},status=400)
        bill=get_object_or_404(Bill,pk=bill_id,is_deleted=False,allocations__line_item__po=po)
        try: payment=create_payment(data={'client':po.client,'received_on':entry_date,'amount':request.data.get('amount'),'created_by':request.user},allocations=[{'bill_id':bill.id,'amount':request.data.get('amount'),'kind':'payment'}])
        except DjangoValidationError as exc:return _validation(exc)
        return Response({'id':payment.id,'kind':'payment'},status=201)
    if kind=='ariba':
        bill_id=request.data.get('bill_id')
        if not bill_id:return Response({'bill_id':['Select the bill to update.']},status=400)
        bill=get_object_or_404(Bill,pk=bill_id,is_deleted=False,allocations__line_item__po=po)
        state=request.data.get('ariba_state')
        if state not in {choice for choice,_ in Bill.Ariba.choices}: return Response({'ariba_state':['Choose a valid Ariba state.']},status=400)
        bill.ariba_state=state; bill.ariba_uploaded_on=entry_date if state in ('uploaded','resubmitted') else None; bill.ariba_reference=request.data.get('reference',''); bill.ariba_rejection_note=request.data.get('note',''); bill.save(update_fields=['ariba_state','ariba_uploaded_on','ariba_reference','ariba_rejection_note','updated_at'])
        return Response({'id':bill.id,'kind':'ariba','ariba_state':bill.ariba_state},status=201)
    line=get_object_or_404(POLineItem,pk=request.data.get('line_item_id'),po=po,is_deleted=False)
    if kind=='delivery':
        if not number:return Response({'number':['Challan number is required.']},status=400)
        try: challan=create_challan(data={'challan_number':number,'challan_date':entry_date,'site':po.site,'created_by':request.user},allocations=[{'line_item_id':line.id,'qty':request.data.get('qty')}])
        except DjangoValidationError as exc:return _validation(exc)
        return Response({'id':challan.id,'kind':'delivery'},status=201)
    if kind=='bill':
        if not number:return Response({'number':['Bill number is required.']},status=400)
        entity,_=LegalEntity.objects.get_or_create(invoice_prefix=number.split('/')[0].upper()[:10] or 'UP',defaults={'name':'Default billing entity'})
        needs_review=False
        try: validate_bill_number(number)
        except DjangoValidationError: needs_review=True
        allocation={'line_item_id':line.id,'qty':request.data.get('qty'),'rate':request.data.get('rate') or line.rate,'gst_rate':request.data.get('gst_rate') or line.gst_rate}
        try:
            bill=Bill.objects.filter(legal_entity=entity,bill_number=number,is_deleted=False).first()
            if bill:
                if bill.allocations.filter(line_item=line).exists():
                    return Response({'number':['This bill is already allocated to this item. Edit the existing record instead.']},status=400)
                allocate_bill(bill=bill,allocations=[allocation])
            else:
                bill=create_bill(data={'legal_entity':entity,'bill_number':number,'bill_date':entry_date,'created_by':request.user,'ariba_state':request.data.get('ariba_state','pending'),'needs_review':needs_review},allocations=[allocation],validate_number=False)
        except DjangoValidationError as exc:return _validation(exc)
        return Response({'id':bill.id,'kind':'bill'},status=201)
    return Response({'kind':['Choose delivery, bill or payment.']},status=400)
@api_view(['POST'])
@permission_classes([CanEditPO])
def po_revise(request,pk):
    po=get_object_or_404(PurchaseOrder,pk=pk,is_deleted=False)
    try: successor=revise_po(po=po,new_lines=request.data.get('lines',[]),reason=request.data.get('reason',''),actor=request.user)
    except DjangoValidationError as exc:return _validation(exc)
    return Response(POSerializer(successor).data,status=201)
@api_view(['POST'])
@permission_classes([CanEditPO])
def po_cancel(request,pk):
    po=get_object_or_404(PurchaseOrder,pk=pk,is_deleted=False); po.status=PurchaseOrder.Status.CANCELLED; po.notes=request.data.get('reason',''); po.updated_by=request.user; po.save(); return Response(POSerializer(po).data)
@api_view(['POST'])
@permission_classes([CanShortClose])
def short_close(request,pk):
    try: line=short_close_line(line=get_object_or_404(POLineItem,pk=pk,is_deleted=False),reason=request.data.get('reason',''),actor=request.user)
    except DjangoValidationError as exc:return _validation(exc)
    return Response({'id':line.id,'short_closed_on':line.short_closed_on,'short_close_reason':line.short_close_reason})
@api_view(['GET'])
@permission_classes([IsTrackerUser])
def search_view(request): return Response(SimplePOSerializer(search(request.query_params.get('q','')),many=True).data)
@api_view(['GET'])
@permission_classes([IsTrackerUser])
def pending_view(request):
    return Response([{'id':line.id,'po_id':line.po_id,'po_number':line.po.po_number,'client':line.po.client.name,'description':line.description,**line_status(line)} for line in pending_lines()])
@api_view(['GET'])
@permission_classes([IsTrackerUser])
def dashboard_view(request):
    return Response(dashboard_snapshot())
@api_view(['GET','PATCH'])
@permission_classes([IsTrackerUser])
def tracker_settings(request):
    settings,_=TrackerSettings.objects.get_or_create(pk=1)
    if request.method=='PATCH':
        for field in ('stuck_after_days','overdue_after_days'):
            if field in request.data:
                value=request.data[field]
                if value in ('',None): setattr(settings,field,None)
                else:
                    try: setattr(settings,field,int(value))
                    except (TypeError,ValueError): return Response({field:['Enter a whole number of days.']},status=400)
        settings.save()
    return Response({'stuck_after_days':settings.stuck_after_days,'overdue_after_days':settings.overdue_after_days})
@api_view(['GET'])
@permission_classes([IsTrackerUser])
def po_export(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    workbook=Workbook(); sheet=workbook.active; sheet.title='PO tracker'
    headers=['PO number','PO date','Lifecycle','Client','Site','Line','Description','Item type','Qty ordered','Rate','PO amount','GST rate','Delivery qty','Billing qty','Item status','Bill amount','Paid amount','Outstanding amount']
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font=Font(bold=True,color='FFFFFF'); cell.fill=PatternFill('solid',fgColor='16324F')
    for po in PurchaseOrder.objects.filter(is_deleted=False).select_related('client','site').prefetch_related('lines__challan_allocations','lines__bill_allocations'):
        totals=po_totals(po)
        for line in po.lines.filter(is_deleted=False):
            derived=line_status(line)
            sheet.append([po.po_number,po.po_date,po.lifecycle_stage,po.client.name,po.site.name if po.site else '',line.line_no,line.description,line.item_type,line.qty_ordered,line.rate,line.amount,line.gst_rate,derived['qty_delivered'],derived['qty_billed'],derived['status'],'',totals['amount_paid'],totals['outstanding_amount']])
    sheet.freeze_panes='A2'; sheet.auto_filter.ref=sheet.dimensions
    for col in sheet.columns:
        sheet.column_dimensions[col[0].column_letter].width=min(max(max(len(str(cell.value or '')) for cell in col)+2, 12), 40)
    for row in sheet.iter_rows(min_row=2, min_col=10, max_col=18):
        for cell in row: cell.number_format='#,##0.00'
    output=io.BytesIO(); workbook.save(output)
    response=HttpResponse(output.getvalue(),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition']='attachment; filename="purchase-orders.xlsx"'
    return response
@api_view(['POST'])
@permission_classes([CanEditPO])
def mark_work_done(request,pk):
    line=get_object_or_404(POLineItem,pk=pk,is_deleted=False); line.work_done_on=request.data.get('work_done_on') or date.today(); line.save(update_fields=['work_done_on','updated_at']); return Response(LineSerializer(line).data)
@api_view(['POST'])
@permission_classes([CanEditPO])
def mark_approved(request,pk):
    line=get_object_or_404(POLineItem,pk=pk,is_deleted=False)
    if not line.work_done_on:return Response({'detail':'Work must be marked done before approval.'},status=400)
    line.client_approved_on=request.data.get('client_approved_on') or date.today(); line.save(update_fields=['client_approved_on','updated_at']); return Response(LineSerializer(line).data)
@api_view(['GET','POST'])
@permission_classes([CanEditPO])
def challan_list(request):
    if request.method=='GET':return Response(ChallanSerializer(Challan.objects.filter(is_deleted=False).prefetch_related('allocations'),many=True).data)
    serializer=ChallanSerializer(data=request.data); serializer.is_valid(raise_exception=True); data=serializer.validated_data; allocations=data.pop('allocations',[])
    data['created_by']=request.user
    try: challan=create_challan(data=data,allocations=[{'line_item_id':a['line_item'].id,'qty':a['qty']} for a in allocations])
    except DjangoValidationError as exc:return _validation(exc)
    return Response(ChallanSerializer(challan).data,status=201)
@api_view(['GET','POST'])
@permission_classes([CanRecordMoney])
def bill_list(request):
    if request.method=='GET':return Response(BillSerializer(Bill.objects.filter(is_deleted=False).prefetch_related('allocations'),many=True).data)
    serializer=BillSerializer(data=request.data); serializer.is_valid(raise_exception=True); data=serializer.validated_data; allocations=data.pop('allocations',[])
    data['created_by']=request.user
    try: bill=create_bill(data=data,allocations=[{'line_item_id':a['line_item'].id,'qty':a['qty'],'rate':a['rate'],'gst_rate':a['gst_rate']} for a in allocations])
    except DjangoValidationError as exc:return _validation(exc)
    return Response(BillSerializer(bill).data,status=201)
@api_view(['GET','POST'])
@permission_classes([CanRecordMoney])
def payment_list(request):
    if request.method=='GET':return Response(PaymentSerializer(Payment.objects.filter(is_deleted=False).prefetch_related('paymentallocation_set'),many=True).data)
    serializer=PaymentSerializer(data=request.data); serializer.is_valid(raise_exception=True); data=serializer.validated_data; allocations=data.pop('paymentallocation_set',[])
    data['created_by']=request.user
    try: payment=create_payment(data=data,allocations=[{'bill_id':a['bill'].id,'amount':a['amount'],'kind':a.get('kind','payment'),'note':a.get('note','')} for a in allocations])
    except DjangoValidationError as exc:return _validation(exc)
    return Response(PaymentSerializer(payment).data,status=201)
@api_view(['POST'])
@permission_classes([CanEditPO])
def paste_preview(request):
    tsv=request.data.get('tsv','')
    if not tsv.strip(): return Response({'tsv':['Paste at least one row from Excel.']},status=400)
    from decimal import Decimal, InvalidOperation
    from .normalizers import normalize_date, normalize_gst, parse_challan_number, classify_item_type
    lines=[]; delivery_groups={}; bill_groups={}; po_number=''; po_date=None
    def decimal_or_none(value):
        try: return Decimal(str(value).replace(',','').strip())
        except (InvalidOperation, ValueError): return None
    for row_number, raw in enumerate(tsv.replace('\r\n','\n').split('\n'),1):
        cells=[cell.strip() for cell in raw.split('\t')]
        if not any(cells): continue
        # Workbook shape has description in F (index 5); the quick shape is Description, Qty, Unit, Rate.
        workbook_shape=len(cells)>=10
        if workbook_shape:
            if cells[1]: po_number=cells[1]
            po_date_raw=cells[2] if len(cells)>2 else ''
            description=cells[5] if len(cells)>5 else ''; qty=cells[6] if len(cells)>6 else ''; unit=cells[7] if len(cells)>7 else ''; rate=cells[8] if len(cells)>8 else ''; item_type=(cells[4] if len(cells)>4 else '').lower()
        else:
            description=cells[0] if len(cells)>0 else ''; qty=cells[1] if len(cells)>1 else ''; unit=cells[2] if len(cells)>2 else ''; rate=cells[3] if len(cells)>3 else ''; item_type=''
            po_date_raw=''
        if not description or description.lower() in ('total','grand total','gst'): continue
        if po_date_raw and po_date is None:
            po_date,_=normalize_date(po_date_raw)
        if item_type not in ('material','service'):
            item_type=classify_item_type(description)
        qty_value=decimal_or_none(qty); rate_value=decimal_or_none(rate)
        if qty_value is None or rate_value is None: return Response({'errors':[{'row':row_number,'message':'Quantity and rate must be numbers.'}]},status=400)
        line_index=len(lines)
        qty=str(qty_value); rate=str(rate_value)
        lines.append({'description':description,'item_type':item_type,'qty_ordered':qty,'unit':unit or 'Nos','rate':rate,'gst_rate':'0'})
        if workbook_shape:
            delivery_qty=decimal_or_none(cells[10] if len(cells)>10 else '')
            challan_raw=cells[13] if len(cells)>13 else ''
            if delivery_qty and delivery_qty>0 and challan_raw:
                challan_number,challan_date=parse_challan_number(challan_raw)
                challan_date=challan_date or po_date or date.today()
                key=(challan_number,challan_date)
                group=delivery_groups.setdefault(key,{'needs_review':not bool(parse_challan_number(challan_raw)[1]),'items':[]})
                group['items'].append({'line_index':line_index,'qty':str(delivery_qty)})
            bill_qty=decimal_or_none(cells[14] if len(cells)>14 else '')
            bill_number=cells[20] if len(cells)>20 else ''
            if bill_qty and bill_qty>0 and bill_number:
                bill_date,bill_date_issue=normalize_date(cells[21] if len(cells)>21 else '')
                bill_date=bill_date or po_date or date.today()
                gst_rate,gst_issue=normalize_gst(cells[17] if len(cells)>17 and cells[17] else '0')
                bill_rate=decimal_or_none(cells[15] if len(cells)>15 else '') or rate_value
                ariba_raw=(cells[22] if len(cells)>22 else '').lower()
                ariba_state='rejected' if 'reject' in ariba_raw else 'resubmitted' if 'resubmit' in ariba_raw else 'uploaded' if 'upload' in ariba_raw else 'pending'
                group=bill_groups.setdefault(bill_number,{'date':bill_date,'ariba_state':ariba_state,'needs_review':bool(bill_date_issue or gst_issue),'items':[]})
                group['items'].append({'line_index':line_index,'qty':str(bill_qty),'rate':str(bill_rate),'gst_rate':str(gst_rate or Decimal('0'))})
    if not lines:return Response({'tsv':['No valid item rows were found.']},status=400)
    if not po_number: po_number=request.data.get('po_number','').strip()
    if not po_number:return Response({'po_number':['The pasted block needs a PO number in column B, or enter one above the grid.']},status=400)
    client_name=request.data.get('client_name','').strip()
    if not client_name:return Response({'client_name':['Choose or enter the client before creating a PO.']},status=400)
    client_code=''.join(ch if ch.isalnum() else '-' for ch in client_name.upper())[:30].strip('-') or 'CLIENT'
    client,_=Client.objects.get_or_create(code=client_code,defaults={'name':client_name})
    site=None; site_name=request.data.get('site_name','').strip()
    if site_name: site,_=Site.objects.get_or_create(client=client,code=site_name[:30],defaults={'name':site_name})
    if PurchaseOrder.objects.filter(client=client,po_number=po_number,is_deleted=False).exists(): return Response({'existing_po':PurchaseOrder.objects.get(client=client,po_number=po_number,is_deleted=False).id,'detail':'This PO already exists. Open it to update it.'},status=409)
    try:
        with transaction.atomic():
            po=create_po(data={'client':client,'site':site,'po_number':po_number,'po_date':po_date,'source':'paste','lifecycle_stage':'draft','lines':lines},actor=request.user)
            created_lines=list(po.lines.order_by('line_no'))
            for (challan_number,challan_date),group in delivery_groups.items():
                challan=Challan.objects.filter(challan_number=challan_number,challan_date=challan_date,is_deleted=False).first()
                allocations=[{'line_item_id':created_lines[item['line_index']].id,'qty':item['qty']} for item in group['items']]
                if challan:
                    for allocation in allocations: ChallanAllocation.objects.get_or_create(challan=challan,line_item_id=allocation['line_item_id'],defaults={'qty':allocation['qty']})
                else:
                    create_challan(data={'challan_number':challan_number,'challan_date':challan_date,'site':site,'source':'paste','needs_review':group['needs_review'],'created_by':request.user},allocations=allocations)
            for bill_number,group in bill_groups.items():
                prefix=bill_number.split('/')[0].upper()[:10] or 'IMPORT'
                entity,_=LegalEntity.objects.get_or_create(invoice_prefix=prefix,defaults={'name':f'{prefix} billing entity'})
                allocations=[{'line_item_id':created_lines[item['line_index']].id,'qty':item['qty'],'rate':item['rate'],'gst_rate':item['gst_rate']} for item in group['items']]
                bill=Bill.objects.filter(legal_entity=entity,bill_number=bill_number,is_deleted=False).first()
                if bill:
                    for allocation in allocations:
                        if not bill.allocations.filter(line_item_id=allocation['line_item_id']).exists(): allocate_bill(bill=bill,allocations=[allocation])
                else:
                    malformed=False
                    try: validate_bill_number(bill_number)
                    except DjangoValidationError: malformed=True
                    bill=create_bill(data={'legal_entity':entity,'bill_number':bill_number,'bill_date':group['date'],'ariba_state':group['ariba_state'],'ariba_uploaded_on':group['date'] if group['ariba_state'] in ('uploaded','resubmitted') else None,'source':'paste','needs_review':group['needs_review'] or malformed,'created_by':request.user},allocations=allocations,validate_number=False)
    except DjangoValidationError as exc:return _validation(exc)
    return Response(POSerializer(po).data,status=201)
@api_view(['POST'])
@permission_classes([CanEditPO])
def pdf_preview(request):
    upload=request.FILES.get('file')
    if not upload:return Response({'detail':'A PDF file is required.'},status=400)
    if not upload.name.lower().endswith('.pdf'):return Response({'detail':'Only PDF files are supported.'},status=400)
    try:
        from pypdf import PdfReader
        text='\n'.join(page.extract_text() or '' for page in PdfReader(upload).pages)
    except Exception:return Response({'detail':'The PDF could not be read. Nothing was saved.'},status=400)
    if not text.strip(): return Response({'detail':'No extractable text found. Nothing was saved.'},status=400)
    return Response({'saved':False,'source_text':text[:20000],'fields':{'po_number':{'value':'','confidence':0.0},'po_date':{'value':'','confidence':0.0}},'lines':[]})
@api_view(['GET'])
@permission_classes([IsAdmin])
def review_list(request):
    qs=ImportReviewItem.objects.all().order_by('-id')
    if request.query_params.get('status')=='open':qs=qs.filter(resolved_at__isnull=True)
    return Response(list(qs.values()))
@api_view(['PATCH'])
@permission_classes([IsAdmin])
def resolve_review(request,pk):
    item=get_object_or_404(ImportReviewItem,pk=pk); item.resolved_at=timezone.now(); item.resolved_by=request.user; item.resolution=request.data.get('resolution',''); item.save(); return Response({'id':item.id,'resolved_at':item.resolved_at})
