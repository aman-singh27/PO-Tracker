"""Import PO Tracker Excel with full lifecycle: POs, Lines, Bills."""
import re
from collections import Counter
from decimal import Decimal, InvalidOperation
from pathlib import Path
from django.db.models import Sum

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook
from datetime import datetime

from tracker.models import (
    Bill, BillAllocation, Client, ImportBatch, ImportReviewItem, 
    LegalEntity, POLineItem, PurchaseOrder, Site,
)
from tracker.normalizers import classify_item_type


SKIP_KEYWORDS = ("grand total", "gst @", "sub total", "total", "total amount")


class Command(BaseCommand):
    help = "Import PO Tracker Excel with POs, lines, bills. --dry-run previews; --commit saves."

    def add_arguments(self, parser):
        parser.add_argument("workbook")
        mode = parser.add_mutually_exclusive_group(required=True)
        mode.add_argument("--dry-run", action="store_true")
        mode.add_argument("--commit", action="store_true")

    def handle(self, *args, **options):
        path = Path(options["workbook"])
        if not path.is_file():
            raise CommandError(f"Workbook not found: {path}")

        records, flags = self._read(path)
        # The same HCL PO can appear in active, pending and working sheets.
        # Dry-run must report the same consolidated set that commit will create.
        consolidated = []
        seen = set()
        for record in records:
            key = (record['client_code'], record['po_number'])
            if key not in seen:
                seen.add(key); consolidated.append(record)
        records = consolidated

        if options["dry_run"]:
            po_count = len(records)
            line_count = sum(len(r["lines"]) for r in records)
            total_value = sum((r["total"] for r in records), Decimal('0'))
            bill_count = sum(len(r.get("bills", {})) for r in records)
            self.stdout.write(str({
                "purchase_orders": po_count,
                "line_items": line_count,
                "bills": bill_count,
                "total_value_inr": str(total_value),
            }))
            return

        # --- commit ---
        seen_po_keys = set()
        with transaction.atomic():
            batch = ImportBatch.objects.create(
                kind="excel",
                filename=path.name,
                rows_total=sum(len(r["lines"]) for r in records),
            )
            imported_pos = 0
            imported_lines = 0
            imported_bills = 0
            flagged = 0

            # Get or create a default legal entity for bills
            entity, _ = LegalEntity.objects.get_or_create(
                invoice_prefix="UP",
                defaults={"name": "Default Entity", "state_code": "09"},
            )

            for record in records:
                po_number = record["po_number"]
                client_code = record["client_code"]
                client_name = record["client_name"]
                po_date = record["po_date"]
                site_code = record.get("site_code")
                site_name = record.get("site_name")
                lines = record["lines"]
                bills_data = record.get("bills", {})

                # Dedupe
                dedupe_key = (client_code, po_number)
                if dedupe_key in seen_po_keys:
                    flagged += 1
                    continue
                seen_po_keys.add(dedupe_key)

                # Get or create client
                client, _ = Client.objects.get_or_create(
                    code=client_code[:30],
                    defaults={"name": client_name[:200]},
                )

                # Get or create site
                site = None
                if site_code:
                    site, _ = Site.objects.get_or_create(
                        client=client,
                        code=site_code[:30],
                        defaults={"name": site_name or client_name[:200]},
                    )

                # Create PO
                po = PurchaseOrder.objects.create(
                    client=client,
                    site=site,
                    po_number=po_number[:100],
                    po_date=po_date,
                    source="migration",
                    needs_review=False,
                )
                imported_pos += 1

                # Create line items and track them for bill allocations
                line_map = {}  # line_no -> POLineItem
                for idx, line_data in enumerate(lines, 1):
                    line_obj = POLineItem.objects.create(
                        po=po,
                        line_no=idx,
                        description=line_data["description"][:500],
                        item_type=line_data["item_type"],
                        qty_ordered=line_data["qty_ordered"],
                        unit=line_data["unit"][:20],
                        rate=line_data["rate"],
                        amount=line_data["amount"],
                        gst_rate=line_data["gst_rate"],
                        source_sheet=line_data.get("source_sheet", ""),
                        source_row=line_data.get("source_row"),
                    )
                    line_map[idx] = line_obj
                    imported_lines += 1

                # Create bills and allocations
                for bill_key, bill_info in bills_data.items():
                    bill_number = bill_info["bill_number"]
                    bill_date = bill_info["bill_date"]
                    
                    # Create or get bill
                    bill, created = Bill.objects.get_or_create(
                        legal_entity=entity,
                        bill_number=bill_number,
                        defaults={
                            "bill_date": bill_date,
                            "basic_amount": Decimal("0"),
                            "gst_amount": Decimal("0"),
                            "total_amount": Decimal("0"),
                        },
                    )
                    if created:
                        imported_bills += 1
                    
                    # Create bill allocations for each line in this bill
                    for alloc in bill_info["allocations"]:
                        line_idx = alloc["line_idx"]
                        if line_idx in line_map:
                            line_obj = line_map[line_idx]
                            BillAllocation.objects.create(
                                bill=bill,
                                line_item=line_obj,
                                qty=alloc["qty"],
                                rate=alloc["rate"],
                                amount=alloc["amount"],
                                gst_rate=alloc["gst_rate"],
                                gst_amount=alloc["gst_amount"],
                                total_amount=alloc["total_amount"],
                            )
                    
                    # Update bill totals
                    totals = bill.allocations.aggregate(
                        basic_sum=Sum('amount'),
                        gst_sum=Sum('gst_amount'),
                        total_sum=Sum('total_amount'),
                    )
                    bill.basic_amount = totals['basic_sum'] or Decimal("0")
                    bill.gst_amount = totals['gst_sum'] or Decimal("0")
                    bill.total_amount = totals['total_sum'] or Decimal("0")
                    bill.save()

            batch.rows_imported = imported_lines
            batch.rows_flagged = flagged
            batch.finished_at = timezone.now()
            batch.save()

        self.stdout.write(self.style.SUCCESS(
            f"Import complete: {imported_pos} POs, {imported_lines} lines, {imported_bills} bills"
        ))

    def _read(self, path):
        """Parse workbook extracting POs, lines, and billing data."""
        workbook = load_workbook(path, data_only=True)
        records = []
        flags = Counter()

        for sheet in workbook.worksheets:
            sheet_title = sheet.title
            if sheet_title == 'Satya Praksh':
                records.extend(self._read_satya_sheet(sheet))
                continue
            if sheet_title == '8100014714':
                record=self._read_standalone_hcl_sheet(sheet)
                if record: records.append(record)
                continue
            current_po = None
            
            for row_no, row in enumerate(sheet.iter_rows(values_only=True), 1):
                values = [self._normalize(v) for v in row]
                
                if not any(values):
                    continue
                
                if self._is_total_row(values):
                    continue
                
                # Columns:
                # 0: S.No, 1: PO No., 2: PO date, 3: Site Address
                # 5: Material Description, 6: Qty, 7: Unit, 8: Rate, 9: Amount
                # 10: Delivery Qty, 13: Delivery Challan No
                # 14: Billing Qty, 15: Billing Rate, 16: Billing Amount
                # 17: GST, 18: GST Amount, 19: Total Billing Amount
                # 20: Bill No, 21: Bill Date
                
                po_number_raw = values[1] if len(values) > 1 else ""
                po_date_raw = values[2] if len(values) > 2 else ""
                site_raw = values[3] if len(values) > 3 else ""
                description = values[5] if len(values) > 5 else ""
                qty_raw = values[6] if len(values) > 6 else ""
                unit_raw = values[7] if len(values) > 7 else ""
                rate_raw = values[8] if len(values) > 8 else ""
                amount_raw = values[9] if len(values) > 9 else ""
                billing_qty_raw = values[14] if len(values) > 14 else ""
                billing_rate_raw = values[15] if len(values) > 15 else ""
                billing_amount_raw = values[16] if len(values) > 16 else ""
                gst_raw = values[17] if len(values) > 17 else ""
                gst_amount_raw = values[18] if len(values) > 18 else ""
                total_billing_raw = values[19] if len(values) > 19 else ""
                bill_no_raw = values[20] if len(values) > 20 else ""
                bill_date_raw = values[21] if len(values) > 21 else ""
                
                # Skip header rows
                if str(po_number_raw).lower() in ("po no.", "po no", "po number"):
                    continue
                
                # New PO. Formats differ by client; only recognised workbook markers
                # are excluded. A valid PO must have a real item description.
                marker = str(po_number_raw).strip().lower()
                if po_number_raw and description and marker not in {'po changed','po amended','po - amended','changed','cancelled','po cancelled','order cancelled by client'} and 'cancelled due to' not in marker and 'changed with tax' not in marker and 'revised with tax' not in marker:
                    site_code, site_name = self._parse_site(site_raw)
                    po_date = self._parse_date(po_date_raw)
                    
                    current_po = {
                        "po_number": str(po_number_raw),
                        "po_date": po_date,
                        "client_code": self._client_for_sheet(sheet_title)[0],
                        "client_name": self._client_for_sheet(sheet_title)[1],
                        "site_code": site_code,
                        "site_name": site_name,
                        "lines": [],
                        "bills": {},
                        "total": Decimal("0"),
                    }
                    records.append(current_po)
                
                # LINE ITEM
                if current_po and description:
                    line = self._parse_line(
                        description, qty_raw, unit_raw, rate_raw, 
                        amount_raw, gst_raw, sheet_title, row_no
                    )
                    if line:
                        line_idx = len(current_po["lines"]) + 1
                        line["line_idx"] = line_idx
                        current_po["lines"].append(line)
                        current_po["total"] += line["amount"]
                        
                        # Check for billing data
                        if bill_no_raw and billing_qty_raw:
                            bill_key = str(bill_no_raw)
                            if bill_key not in current_po["bills"]:
                                current_po["bills"][bill_key] = {
                                    "bill_number": bill_no_raw,
                                    "bill_date": self._parse_date(bill_date_raw),
                                    "allocations": [],
                                }
                            
                            # Parse billing values
                            try:
                                billing_qty = Decimal(str(billing_qty_raw).replace(",", ""))
                            except:
                                billing_qty = line["qty_ordered"]
                            
                            try:
                                billing_rate = Decimal(str(billing_rate_raw).replace(",", ""))
                            except:
                                billing_rate = line["rate"]
                            
                            try:
                                billing_amount = Decimal(str(billing_amount_raw).replace(",", ""))
                            except:
                                billing_amount = billing_qty * billing_rate
                            
                            try:
                                gst_rate = Decimal(str(gst_raw).replace(",", ""))
                                if gst_rate > 1:
                                    gst_rate = gst_rate / 100
                            except:
                                gst_rate = Decimal("0")
                            
                            try:
                                gst_amount = Decimal(str(gst_amount_raw).replace(",", ""))
                            except:
                                gst_amount = Decimal("0")
                            
                            try:
                                total_billing = Decimal(str(total_billing_raw).replace(",", ""))
                            except:
                                total_billing = billing_amount + gst_amount
                            
                            current_po["bills"][bill_key]["allocations"].append({
                                "line_idx": line_idx,
                                "qty": billing_qty,
                                "rate": billing_rate,
                                "amount": billing_amount,
                                "gst_rate": gst_rate,
                                "gst_amount": gst_amount,
                                "total_amount": total_billing,
                            })

        return records, flags

    def _client_for_sheet(self, title):
        lowered=title.lower()
        if title in {'HCL PO','Pending PO','Sheet1','HCL Sec-60','8100014714'}: return ('HCL','HCL Technologies')
        if 'dlf' in lowered: return ('DLF','DLF Mall of India')
        if 'metlife' in lowered: return ('METLIFE','Metlife')
        if 'satya' in lowered: return ('SATYA-PRAKASH','Satya Prakash')
        return (title.upper()[:30].replace(' ','-').replace('.','-'),title)

    def _read_satya_sheet(self, sheet):
        records=[]; client_code,client_name=self._client_for_sheet(sheet.title)
        for row_no,row in enumerate(sheet.iter_rows(values_only=True),1):
            values=[self._normalize(value) for value in row]
            po_number=values[1] if len(values)>1 else ''
            description=values[7] if len(values)>7 else ''
            basic=values[4] if len(values)>4 else ''
            if not po_number or str(po_number).lower() in {'po no','po no.'} or not description: continue
            try: amount=Decimal(str(basic).replace(',',''))
            except (InvalidOperation,ValueError): continue
            records.append({'po_number':str(po_number),'po_date':self._parse_date(values[2] if len(values)>2 else ''),'client_code':client_code,'client_name':client_name,'site_code':(values[3] if len(values)>3 else '')[:30] or None,'site_name':values[3] if len(values)>3 else '', 'lines':[{'description':description,'item_type':classify_item_type(description),'qty_ordered':Decimal('1'),'unit':'Job','rate':amount,'amount':amount,'gst_rate':Decimal('0'),'source_sheet':sheet.title,'source_row':row_no}], 'bills':{},'total':amount})
        return records

    def _read_standalone_hcl_sheet(self, sheet):
        client_code,client_name=self._client_for_sheet(sheet.title); po_number='8100014714'; po_date=None; site_code=None; site_name=''; lines=[]
        for row_no,row in enumerate(sheet.iter_rows(values_only=True),1):
            values=[self._normalize(value) for value in row]
            text=' '.join(values[:2]).lower()
            if 'po. no' in text:
                po_number=values[1].split('-')[-1].strip()
            elif 'po date' in text:
                po_date=self._parse_date(values[1].split('-')[-1].strip())
            elif 'site-' in text:
                site_code,site_name=self._parse_site(values[1].split('-',1)[-1].strip())
            # This sheet shifts description/qty/unit/rate/amount one column left.
            if len(values)>5 and values[1] and values[1].lower() not in {'material description','total','grand total'}:
                try: qty=Decimal(values[2].replace(',','')); rate=Decimal(values[4].replace(',','')); amount=Decimal(values[5].replace(',',''))
                except (InvalidOperation,ValueError): continue
                lines.append({'description':values[1],'item_type':classify_item_type(values[1]),'qty_ordered':qty,'unit':values[3] or 'Nos','rate':rate,'amount':amount,'gst_rate':Decimal('0'),'source_sheet':sheet.title,'source_row':row_no})
        if not lines:return None
        return {'po_number':po_number,'po_date':po_date,'client_code':client_code,'client_name':client_name,'site_code':site_code,'site_name':site_name,'lines':lines,'bills':{},'total':sum((line['amount'] for line in lines),Decimal('0'))}

    def _normalize(self, value):
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%d/%m/%Y")
        return str(value).strip()

    def _is_total_row(self, values):
        text = " ".join(str(v).lower() for v in values if v)
        return any(kw in text for kw in SKIP_KEYWORDS)

    def _parse_site(self, site_raw):
        if not site_raw:
            return None, None
        match = re.match(r"\((\w+)\)\s*(.+)", site_raw)
        if match:
            return match.group(1), match.group(1)
        return site_raw[:30], site_raw[:30]

    def _parse_date(self, date_raw):
        if not date_raw:
            return None
        for fmt in ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d-%b-%Y"]:
            try:
                return datetime.strptime(str(date_raw), fmt).date()
            except (ValueError, TypeError):
                continue
        if isinstance(date_raw, datetime):
            return date_raw.date()
        return None

    def _parse_line(self, description, qty_raw, unit_raw, rate_raw, amount_raw, gst_raw, sheet_title, row_no):
        if not description:
            return None
        
        desc_lower = description.lower()
        if any(kw in desc_lower for kw in SKIP_KEYWORDS):
            return None
        
        try:
            qty = Decimal(str(qty_raw).replace(",", ""))
            if qty <= 0:
                qty = Decimal("1")
        except (InvalidOperation, ValueError):
            qty = Decimal("1")
        
        try:
            rate = Decimal(str(rate_raw).replace(",", ""))
        except (InvalidOperation, ValueError):
            rate = Decimal("0")
        
        try:
            amount = Decimal(str(amount_raw).replace(",", ""))
        except (InvalidOperation, ValueError):
            amount = (qty * rate).quantize(Decimal("0.01"))
        
        try:
            gst_rate = Decimal(str(gst_raw).replace(",", ""))
            if gst_rate > 1:
                gst_rate = gst_rate / 100
        except (InvalidOperation, ValueError):
            gst_rate = Decimal("0")
        
        return {
            "description": description,
            "item_type": classify_item_type(description),
            "qty_ordered": qty,
            "unit": unit_raw or "Nos",
            "rate": rate,
            "amount": amount,
            "gst_rate": gst_rate,
            "source_sheet": sheet_title,
            "source_row": row_no,
        }
