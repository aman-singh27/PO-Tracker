"""Simple import focusing on POs first, bills later."""
from pathlib import Path
from decimal import Decimal
from datetime import datetime

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from tracker.models import (
    Client, ImportBatch, ImportReviewItem, POLineItem, PurchaseOrder, Site,
)
from tracker.normalizers import classify_item_type


class Command(BaseCommand):
    help = "Import POs from Excel"

    def add_arguments(self, parser):
        parser.add_argument("workbook")
        mode = parser.add_mutually_exclusive_group(required=True)
        mode.add_argument("--dry-run", action="store_true")
        mode.add_argument("--commit", action="store_true")

    def handle(self, *args, **options):
        path = Path(options["workbook"])
        if not path.is_file():
            raise CommandError(f"Workbook not found: {path}")

        records = self._read(path)

        if options["dry_run"]:
            po_count = len(records)
            line_count = sum(len(r["lines"]) for r in records)
            total_value = sum(r["total"] for r in records)
            self.stdout.write(str({
                "purchase_orders": po_count,
                "line_items": line_count,
                "total_value_inr": str(total_value),
            }))
            return

        # Import
        with transaction.atomic():
            batch = ImportBatch.objects.create(
                kind="excel",
                filename=path.name,
                rows_total=sum(len(r["lines"]) for r in records),
            )
            imported_pos = 0
            imported_lines = 0

            for record in records:
                # Get or create client
                client, _ = Client.objects.get_or_create(
                    code=record["client_code"],
                    defaults={"name": record["client_name"]},
                )

                # Get or create site
                site = None
                if record.get("site_code"):
                    site, _ = Site.objects.get_or_create(
                        client=client,
                        code=record["site_code"],
                        defaults={"name": record.get("site_name", "")},
                    )

                # Check if PO already exists for this client
                if PurchaseOrder.objects.filter(client=client, po_number=record["po_number"], is_deleted=False).exists():
                    ImportReviewItem.objects.create(
                        batch=batch,
                        severity="warning",
                        reason_code="DUPLICATE_PO",
                        source_ref=f"{record['client_code']}/{record['po_number']}",
                        payload_json={"client": record["client_code"], "po_number": record["po_number"]},
                    )
                    continue

                # Create PO
                po = PurchaseOrder.objects.create(
                    client=client,
                    site=site,
                    po_number=record["po_number"],
                    po_date=record["po_date"],
                    source="migration",
                    needs_review=False,
                )
                imported_pos += 1

                # Create line items
                for idx, line_data in enumerate(record["lines"], 1):
                    POLineItem.objects.create(
                        po=po,
                        line_no=idx,
                        description=line_data["description"][:500],
                        item_type=line_data["item_type"],
                        qty_ordered=line_data["qty_ordered"],
                        unit=line_data["unit"],
                        rate=line_data["rate"],
                        amount=line_data["amount"],
                        gst_rate=line_data["gst_rate"],
                        source_sheet=line_data.get("source_sheet", ""),
                        source_row=line_data.get("source_row"),
                    )
                    imported_lines += 1

            batch.rows_imported = imported_lines
            batch.finished_at = timezone.now()
            batch.save()

        self.stdout.write(self.style.SUCCESS(
            f"Import complete: {imported_pos} POs, {imported_lines} lines"
        ))

    def _read(self, path):
        """Simple parser for POs only."""
        workbook = load_workbook(path, data_only=True)
        records = []

        for sheet in workbook.worksheets:
            sheet_title = sheet.title
            current_po = None
            
            for row_no, row in enumerate(sheet.iter_rows(values_only=True), 1):
                values = [self._normalize(v) for v in row]
                
                # Skip empty rows
                if not any(values):
                    continue
                
                # Skip header/total rows
                if self._is_total_row(values):
                    continue
                
                # Columns
                po_number_raw = values[1] if len(values) > 1 else ""
                po_date_raw = values[2] if len(values) > 2 else ""
                site_raw = values[3] if len(values) > 3 else ""
                description = values[5] if len(values) > 5 else ""
                qty_raw = values[6] if len(values) > 6 else ""
                unit_raw = values[7] if len(values) > 7 else ""
                rate_raw = values[8] if len(values) > 8 else ""
                amount_raw = values[9] if len(values) > 9 else ""
                
                # Skip header rows
                if str(po_number_raw).lower() in ("po no.", "po no", "po number"):
                    continue
                
                # New PO
                if po_number_raw and str(po_number_raw).strip().replace(".", "").isdigit():
                    site_code, site_name = self._parse_site(site_raw)
                    po_date = self._parse_date(po_date_raw)
                    
                    current_po = {
                        "po_number": str(po_number_raw).strip(),
                        "po_date": po_date,
                        "client_code": sheet_title.upper()[:30].replace(" ", "-").replace(".", "-"),
                        "client_name": sheet_title,
                        "site_code": site_code,
                        "site_name": site_name,
                        "lines": [],
                        "total": Decimal("0"),
                    }
                    records.append(current_po)
                
                # Line item
                if current_po and description and description.lower() not in ("total", "gst"):
                    line = self._parse_line(
                        description, qty_raw, unit_raw, rate_raw, amount_raw, sheet_title, row_no
                    )
                    if line:
                        current_po["lines"].append(line)
                        current_po["total"] += line["amount"]

        return records

    def _normalize(self, value):
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%d/%m/%Y")
        return str(value).strip()

    def _is_total_row(self, values):
        text = " ".join(str(v).lower() for v in values if v)
        keywords = ["grand total", "gst @", "sub total", "total", "gst", "grand"]
        return any(kw in text for kw in keywords)

    def _parse_site(self, site_raw):
        if not site_raw:
            return None, None
        import re
        match = re.match(r"\((\w+)\)\s*(.+)", site_raw)
        if match:
            return match.group(1), match.group(1)
        return site_raw[:30], site_raw[:30]

    def _parse_date(self, date_raw):
        if not date_raw:
            return None
        for fmt in ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"]:
            try:
                return datetime.strptime(str(date_raw), fmt).date()
            except (ValueError, TypeError):
                continue
        if isinstance(date_raw, datetime):
            return date_raw.date()
        return None

    def _parse_line(self, description, qty_raw, unit_raw, rate_raw, amount_raw, sheet_title, row_no):
        if not description:
            return None
        
        desc_lower = description.lower()
        if any(kw in desc_lower for kw in ["total", "gst", "grand"]):
            return None
        
        try:
            qty = Decimal(str(qty_raw).replace(",", ""))
            if qty <= 0:
                qty = Decimal("1")
        except:
            qty = Decimal("1")
        
        try:
            rate = Decimal(str(rate_raw).replace(",", ""))
        except:
            rate = Decimal("0")
        
        try:
            amount = Decimal(str(amount_raw).replace(",", ""))
        except:
            amount = (qty * rate).quantize(Decimal("0.01"))
        
        return {
            "description": description[:500],
            "item_type": classify_item_type(description),
            "qty_ordered": qty,
            "unit": unit_raw or "Nos",
            "rate": rate,
            "amount": amount,
            "gst_rate": Decimal("0.18"),  # Default GST
            "source_sheet": sheet_title,
            "source_row": row_no,
        }
